import csv
import io
from datetime import date
from decimal import Decimal, InvalidOperation

from django import forms
from django.conf import settings
from django.contrib import messages
from django.db.models import Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_http_methods, require_POST
from openpyxl import load_workbook

from accounts.permissions import require_admin
from core.modal import modal_form, modal_success
from finanzas.models import CuentaBancaria, DocumentoSII, MovimientoCuenta, PeriodoTributario
from presupuestos.models import FacturaBoleta


def _recalcular_periodo(year: int, month: int) -> PeriodoTributario:
    """IVA/PPM desde FacturaBoleta; DocumentoSII queda asociado al período como soporte."""
    periodo = date(year, month, 1)
    agg = FacturaBoleta.objects.filter(fecha__year=year, fecha__month=month).aggregate(
        total_neto=Sum('monto_neto'),
        iva_debito=Sum('monto_iva'),
    )
    total_neto = agg['total_neto'] or Decimal('0')
    iva_debito = agg['iva_debito'] or Decimal('0')
    pt, _ = PeriodoTributario.objects.get_or_create(
        periodo=periodo,
        defaults={'iva_credito': Decimal('0')},
    )
    iva_credito = pt.iva_credito
    iva_a_pagar = max(iva_debito - iva_credito, Decimal('0'))
    ppm_rate = Decimal(str(settings.FINANZAS_PPM_RATE))
    pt.total_neto = total_neto
    pt.iva_debito = iva_debito
    pt.iva_a_pagar = iva_a_pagar
    pt.ppm = (total_neto * ppm_rate).quantize(Decimal('0.01'))
    pt.save()
    return pt


def _row_get(row: dict, *keys):
    for key in keys:
        for k, v in row.items():
            if str(k or '').strip().lower() == key:
                return v
    return None


def _parse_factura_id(row: dict):
    raw = _row_get(row, 'factura_id', 'factura')
    if raw is None or str(raw).strip() == '':
        return None
    try:
        return int(str(raw).strip())
    except (ValueError, TypeError):
        return None


def _parse_movimiento_row(row: dict) -> MovimientoCuenta | None:
    fecha_raw = _row_get(row, 'fecha')
    if fecha_raw is None or str(fecha_raw).strip() == '':
        return None
    try:
        if isinstance(fecha_raw, date):
            fecha = fecha_raw
        elif '/' in str(fecha_raw):
            d, m, y = str(fecha_raw).split('/')[:3]
            fecha = date(int(y), int(m), int(d))
        else:
            fecha = date.fromisoformat(str(fecha_raw)[:10])
    except (ValueError, TypeError):
        return None
    tipo = str(_row_get(row, 'tipo') or '').strip().lower()
    if tipo not in (MovimientoCuenta.CARGO, MovimientoCuenta.ABONO):
        return None
    try:
        monto = Decimal(str(_row_get(row, 'monto') or '0'))
    except InvalidOperation:
        return None

    cuenta = str(_row_get(row, 'cuenta') or 'principal').strip() or 'principal'
    factura_id = _parse_factura_id(row)
    factura = None
    if factura_id is not None:
        factura = FacturaBoleta.objects.filter(pk=factura_id).first()

    return MovimientoCuenta(
        fecha=fecha,
        cuenta=cuenta,
        tipo=tipo,
        monto=monto,
        descripcion=str(_row_get(row, 'descripcion') or '').strip(),
        referencia=str(_row_get(row, 'referencia') or '').strip(),
        factura=factura,
    )


def _import_movimientos(rows) -> int:
    objs = []
    for row in rows:
        mov = _parse_movimiento_row(row)
        if mov:
            objs.append(mov)
    if objs:
        MovimientoCuenta.objects.bulk_create(objs)
    return len(objs)


def _form_context(**extra):
    ctx = {
        'cuentas': CuentaBancaria.objects.all(),
        'facturas': FacturaBoleta.objects.select_related('proyecto').all()[:200],
    }
    ctx.update(extra)
    return ctx


@require_admin
def dashboard(request):
    periodos = PeriodoTributario.objects.all()[:12]
    ingresos = MovimientoCuenta.objects.filter(tipo=MovimientoCuenta.ABONO).aggregate(
        t=Sum('monto')
    )['t'] or Decimal('0')
    egresos = MovimientoCuenta.objects.filter(tipo=MovimientoCuenta.CARGO).aggregate(
        t=Sum('monto')
    )['t'] or Decimal('0')
    recientes = MovimientoCuenta.objects.select_related(
        'cuenta_bancaria', 'factura'
    ).all()[:15]
    return render(
        request,
        'finanzas/dashboard.html',
        {
            'periodos': periodos,
            'ingresos': ingresos,
            'egresos': egresos,
            'balance': ingresos - egresos,
            'movimientos_recientes': recientes,
        },
    )


@require_admin
def periodos_list(request):
    periodos = PeriodoTributario.objects.all()
    return render(request, 'finanzas/periodos.html', {'periodos': periodos})


@require_admin
@require_POST
def recalcular_periodo(request, year, month):
    if not (1 <= month <= 12):
        messages.error(request, 'Mes no válido.')
        return redirect('finanzas:periodos_list')
    pt = _recalcular_periodo(year, month)
    messages.success(request, f'Período {pt.periodo:%Y-%m} recalculado.')
    return redirect('finanzas:periodos_list')


@require_admin
def documentos_sii(request):
    docs = DocumentoSII.objects.select_related('uploaded_by').all()
    return render(request, 'finanzas/sii_list.html', {'documentos': docs})


@require_admin
@require_http_methods(['GET', 'POST'])
def documento_sii_upload(request):
    list_url = reverse('finanzas:sii_list')
    form = forms.Form()  # ponytail: body HTML en template; sin ModelForm
    if request.method == 'POST':
        titulo = (request.POST.get('titulo') or '').strip()
        periodo_raw = request.POST.get('periodo')
        archivo = request.FILES.get('archivo')
        notas = (request.POST.get('notas') or '').strip()
        if not titulo or not periodo_raw or not archivo:
            messages.error(request, 'Título, período y archivo son obligatorios.')
        else:
            try:
                p = date.fromisoformat(periodo_raw)
                periodo = date(p.year, p.month, 1)
            except ValueError:
                messages.error(request, 'Fecha de período no válida.')
                return modal_form(
                    request,
                    title='Subir documento SII',
                    form=form,
                    action_url=reverse('finanzas:sii_upload'),
                    multipart=True,
                    extra={
                        'cancel_url': list_url,
                        'form_body_template': 'finanzas/sii_form_body.html',
                    },
                )
            DocumentoSII.objects.create(
                titulo=titulo,
                periodo=periodo,
                archivo=archivo,
                notas=notas,
                uploaded_by=request.user,
            )
            messages.success(request, 'Documento SII subido.')
            return modal_success(request, list_url)
    return modal_form(
        request,
        title='Subir documento SII',
        form=form,
        action_url=reverse('finanzas:sii_upload'),
        multipart=True,
        extra={
            'cancel_url': list_url,
            'form_body_template': 'finanzas/sii_form_body.html',
        },
    )


@require_admin
@require_POST
def documento_sii_delete(request, pk):
    doc = get_object_or_404(DocumentoSII, pk=pk)
    if doc.archivo:
        doc.archivo.delete(save=False)
    doc.delete()
    messages.success(request, 'Documento eliminado.')
    return redirect('finanzas:sii_list')


@require_admin
def movimientos_list(request):
    movimientos = MovimientoCuenta.objects.select_related(
        'cuenta_bancaria', 'factura'
    ).all()[:200]
    return render(request, 'finanzas/movimientos.html', {'movimientos': movimientos})


@require_admin
@require_http_methods(['GET', 'POST'])
def movimiento_create(request):
    list_url = reverse('finanzas:movimientos_list')
    form = forms.Form()  # ponytail: body HTML en template; sin ModelForm
    action_url = reverse('finanzas:movimiento_create')

    def _modal():
        return modal_form(
            request,
            title='Nuevo movimiento',
            form=form,
            action_url=action_url,
            extra=_form_context(
                cancel_url=list_url,
                form_body_template='finanzas/movimiento_form_body.html',
            ),
        )

    if request.method == 'POST':
        fecha_raw = request.POST.get('fecha')
        tipo = request.POST.get('tipo')
        monto_raw = request.POST.get('monto')
        try:
            fecha = date.fromisoformat(fecha_raw)
            monto = Decimal(monto_raw)
        except (ValueError, TypeError, InvalidOperation):
            messages.error(request, 'Datos no válidos.')
            return _modal()
        if tipo not in (MovimientoCuenta.CARGO, MovimientoCuenta.ABONO):
            messages.error(request, 'Tipo no válido.')
            return _modal()

        cuenta_bancaria = None
        cb_id = (request.POST.get('cuenta_bancaria') or '').strip()
        if cb_id:
            cuenta_bancaria = CuentaBancaria.objects.filter(pk=cb_id).first()

        factura = None
        factura_id = (request.POST.get('factura') or '').strip()
        if factura_id:
            factura = FacturaBoleta.objects.filter(pk=factura_id).first()

        cuenta = (request.POST.get('cuenta') or '').strip()
        if cuenta_bancaria and not cuenta:
            cuenta = cuenta_bancaria.nombre
        if not cuenta:
            cuenta = 'principal'

        MovimientoCuenta.objects.create(
            fecha=fecha,
            cuenta=cuenta,
            cuenta_bancaria=cuenta_bancaria,
            factura=factura,
            tipo=tipo,
            monto=monto,
            descripcion=(request.POST.get('descripcion') or '').strip(),
            referencia=(request.POST.get('referencia') or '').strip(),
        )
        messages.success(request, 'Movimiento registrado.')
        return modal_success(request, list_url)
    return _modal()


@require_admin
@require_http_methods(['GET', 'POST'])
def movimientos_import(request):
    if request.method == 'POST':
        fmt = request.POST.get('formato', 'csv')
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Selecciona un archivo.')
            return render(request, 'finanzas/import.html')
        try:
            if fmt == 'xlsx':
                wb = load_workbook(archivo, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(max_row=1))]
                rows = [
                    {headers[i]: (cell.value if cell.value is not None else '') for i, cell in enumerate(row)}
                    for row in ws.iter_rows(min_row=2)
                ]
            else:
                text = io.TextIOWrapper(archivo.file, encoding='utf-8-sig')
                rows = list(csv.DictReader(text))
            n = _import_movimientos(rows)
            messages.success(request, f'{n} movimiento(s) importado(s).')
            return redirect('finanzas:movimientos_list')
        except Exception:
            messages.error(request, 'No se pudo leer el archivo.')
    return render(request, 'finanzas/import.html')
